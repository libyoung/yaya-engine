#!/usr/bin/env python
# -*- coding:utf-8 -*-
##############################
#   Create Time: 20170815
#   Author:      liuyang
#   Email:       czyang.liu@jrdcom.com
#   Content:     Create xlsx file
##############################
import openpyxl
from openpyxl.styles import colors, PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import Worksheet
from logdatatojson import LJson
import datetime
#################################################
#   Can Custom Contents
#################################################
SHEET1_NAME = "Stability Time Per Section"
SHEET2_NAME = "Completion Percentage"
ALL_SECTION_LIST= [ "Telephony",
                    "Messaging",
                    "Email",
                    "Browser",
                    "Storefront",
                    "PIM",
                    "MultiMedia",
                    "MultiTasking",
                    "MenuNavigation",
                    "WiFi",
                    "VoLTE",
                    "SMSOverIP",
                    "WFC"
                    ]
MAX_SECTION_LIST = []
DEVICE_LIST = []
DEVICENUM = 7
FILENAME="StabilityTest_Report.xlsx"


#################################################
#   FONT
#################################################
ARIAL_10_FONT = Font(name='Arial',size=10)
TAHOMA_8_FONT = Font(name='Tahoma',size=8)
TAHOMA_8_BOLD_FONT = Font(name='Tahoma', size=8, bold = True)

#################################################
#   ALIGNMENT
#################################################
CENTER_CENTER_ALIGNMENT =  Alignment(horizontal='center', vertical='center')
LEFT_CENTER_ALIGNMENT =  Alignment(horizontal='left', vertical='center')
RIGHT_CENTER_ALIGNMENT =  Alignment(horizontal='right', vertical='center')

#################################################
#   FILL: fill color
#################################################
LOOP_BG_COLOR_FILL= PatternFill(fill_type = 'solid', start_color = "00ff99db")
SECTION_BG_COLOR_FILL= PatternFill(fill_type = 'solid', start_color = "00ffff99")
SECTION_AVG_TIME_BG_COLOR_FILL = PatternFill(fill_type = 'solid', start_color = "00ffcc99")
LOOP_TIME_BG_COLOR_FILL = PatternFill(fill_type = 'solid', start_color = "00ccffff")
TITLE_BG_COLOR_FILL = PatternFill(fill_type = 'solid', start_color = "00ccffcc")
SHEET_END_BG_COLOR_FILL = PatternFill(fill_type = 'solid', start_color = "0099ccff")
SUMMARY_BG_COLOR_FILL = PatternFill(fill_type = 'solid', start_color = "00fcf304")

#################################################
#   BORDER
#################################################
TOP_THIN_BORDER = Border(top=Side(border_style='thin',color='00000000'))
BOTTOM_THIN_BORDER = Border(bottom=Side(border_style='thin',color='00000000'))
LEFT_THIN_BORDER = Border(left=Side(border_style='thin',color='00000000'))
RIGHT_THIN_BORDER = Border(right=Side(border_style='thin',color='00000000'))

#################################################
#   SET ROW COLUMN width High, Freeze panes
#################################################
wb = openpyxl.Workbook()
sheet1 = wb.active
sheet1.title = SHEET1_NAME
sheet2 = wb.create_sheet(title=SHEET2_NAME)

sheet1.column_dimensions['A'].width = 20
sheet1.column_dimensions['V'].width = 20
#sheet1.freeze_panes = "A25"

sheet2.column_dimensions['B'].width = 20

sheet1.row_dimensions.height = 11

#################################################
#   Create the Area four side_border
#################################################
def set_area_four_side_border(sheet, area, style = 'medium',color='00000000'):
    top_border = Border(top=Side(border_style=style,color=color))
    bottom_border = Border(bottom=Side(border_style=style,color=color))
    left_border = Border(left=Side(border_style=style,color=color))
    right_border = Border(right=Side(border_style=style,color=color))

    rows=sheet[area]
    if len(rows)==1:
        for cell in rows[0][1:-1]:
            cell.border = top_border + bottom_border
        rows[0][0].border =top_border + bottom_border+left_border
        rows[0][-1].border =top_border + bottom_border+right_border
    else:
        for cell in rows[0][1:-1]:
            cell.border = top_border

        for cell in rows[-1][1:-1]:
            cell.border = bottom_border

        for row in rows[1:-1]:
            row[0].border = left_border
            row[-1].border = right_border

        rows[0][0].border = top_border+left_border
        rows[0][-1].border = top_border+right_border
        rows[-1][0].border = bottom_border +left_border
        rows[-1][-1].border = bottom_border + right_border

def set_area_all_cell_four_side_border(sheet, area, style = 'thin',color='00000000'):
    rows=sheet[area]
    for row in rows:
        for cell in row:
            cell.border=get_area_four_side_border(style=style,color=color)

def set_area_all_cell(sheet, area, style = 'thin',color='00000000',value='0:00:00',font = TAHOMA_8_FONT):
    rows=sheet[area]
    for row in rows:
        for cell in row:
            cell.value = value
            cell.number_format = "[h]:mm:ss"
            cell.font = font
            cell.alignment = RIGHT_CENTER_ALIGNMENT
            cell.border=get_area_four_side_border(style=style,color=color)

def set_area_all_cell_2(sheet, area, style = 'thin',color='00000000',font=TAHOMA_8_FONT):
    rows=sheet[area]
    for row in rows:
        for cell in row:
            cell.font = font
            cell.alignment = RIGHT_CENTER_ALIGNMENT
            cell.border=get_area_four_side_border(style=style,color=color)

def get_area_four_side_border(style = 'thin',color='00000000'):
    top_border = Border(top=Side(border_style=style,color=color))
    bottom_border = Border(bottom=Side(border_style=style,color=color))
    left_border = Border(left=Side(border_style=style,color=color))
    right_border = Border(right=Side(border_style=style,color=color))
    return top_border+bottom_border+left_border+right_border

#################################################
#   Create Rights area 
#################################################
def set_Rights_area(sheet):
    RIGHTS_STR = "© 2011 AT&T Intellectual Property. All rights reserved. AT&T, AT&T logo and all other marks contained herein are trademarks of AT&T Intellectual Property and/or AT&T affiliated companies."
    RIGHTS_AREA = "A3:V7"

    sheet.merge_cells(RIGHTS_AREA)
    Rights_cell = sheet[RIGHTS_AREA.split(":")[0]]
    Rights_cell.value = RIGHTS_STR
    Rights_cell.fill = SHEET_END_BG_COLOR_FILL
    Rights_cell.font = ARIAL_10_FONT
    Rights_cell.alignment=CENTER_CENTER_ALIGNMENT

    set_area_four_side_border(sheet,RIGHTS_AREA)
#################################################
#   Create Table Instructions area
#################################################
def set_Table_instructions_area(sheet, info):
    TABLE_INSTRCUCTIONS_BG_COLOR = "00ccffff" #(204,255,255)
    TABLE_INSTRCUCTIONS_AREA = 'B10:N15'

    for row, entry in enumerate(info, start=0):
        sheet.cell(row=10+row, column=2, value=entry).font = ARIAL_10_FONT

    fill = PatternFill(fill_type = 'solid', start_color = TABLE_INSTRCUCTIONS_BG_COLOR)

    for row in sheet[TABLE_INSTRCUCTIONS_AREA]:
        for cell in row:
            cell.fill = fill

    set_area_four_side_border(sheet, TABLE_INSTRCUCTIONS_AREA)

#################################################
#   Create Summary area
#################################################
def set_Loop_time_summary_area():

    for row in sheet1["B19:B21"]:
        row[0].fill = SUMMARY_BG_COLOR_FILL

    for col in xrange(1,DEVICENUM+1):
        title_cell = sheet1.cell(row=19,column=col+2)
        title_cell.fill = SUMMARY_BG_COLOR_FILL

        cell=sheet1.cell(row=20, column=col+2, value="Device"+str(col))
        cell.font=TAHOMA_8_FONT
        cell.alignment=CENTER_CENTER_ALIGNMENT

        value_cell = sheet1.cell(row=21, column=col+2)
        value_cell.number_format = "[h]:mm:ss"
        value_cell.font = TAHOMA_8_BOLD_FONT
        value_cell.alignment = CENTER_CENTER_ALIGNMENT

    end_col_letter = get_column_letter(col+2)

    sheet1.merge_cells("C19:%s19" % end_col_letter)
    #must fisrt merge cell and set border
    set_area_four_side_border(sheet1, "B19:%s21" % end_col_letter, style='thin')

    set_area_all_cell_four_side_border(sheet1, "C20:%s21" % end_col_letter)

    summary_title_cell = sheet1['C19'] 
    summary_title_cell.value= "Summary - Average Loop Time Per Device"
    summary_title_cell.font = TAHOMA_8_BOLD_FONT
    summary_title_cell.alignment=CENTER_CENTER_ALIGNMENT

    time_cell = sheet1['B21']
    time_cell.value = "Time"
    time_cell.font = TAHOMA_8_BOLD_FONT
    time_cell.alignment=CENTER_CENTER_ALIGNMENT
    time_cell.border = get_area_four_side_border()

    total_avg_cell = sheet1['K19']
    total_avg_cell.value = 'Total Avg'
    total_avg_cell.fill = SUMMARY_BG_COLOR_FILL
    total_avg_cell.font = TAHOMA_8_BOLD_FONT
    total_avg_cell.alignment=CENTER_CENTER_ALIGNMENT
    total_avg_cell.border = get_area_four_side_border()

    total_avg_value_cell =sheet1['K20']
    total_avg_value_cell.value = '=SUM(C21:%s21)/COUNTIF(C21:%s21,">0")' % (end_col_letter, end_col_letter)
    total_avg_value_cell.font = TAHOMA_8_FONT
    total_avg_value_cell.alignment=CENTER_CENTER_ALIGNMENT
    total_avg_value_cell.border = get_area_four_side_border()
    total_avg_value_cell.number_format = "[h]:mm:ss"


def set_Success_rate_summary_area():
    for col,section in enumerate(MAX_SECTION_LIST,start=0):
        title_cell=sheet2.cell(row=19, column=col+3)
        title_cell.fill = SUMMARY_BG_COLOR_FILL

        cell=sheet2.cell(row=20, column=col+3, value=section)
        cell.font=TAHOMA_8_FONT
        cell.alignment=Alignment(wrap_text=True, horizontal='left', vertical='top')
        cell.border = get_area_four_side_border()

        value_cell=sheet2.cell(row=21, column=col+3)
        value_cell.font = TAHOMA_8_FONT
        value_cell.alignment = CENTER_CENTER_ALIGNMENT
        value_cell.border = get_area_four_side_border()

        requirement_cell=sheet2.cell(row=22, column=col+3, value="95%")
        requirement_cell.font = TAHOMA_8_BOLD_FONT
        requirement_cell.alignment = CENTER_CENTER_ALIGNMENT
        requirement_cell.fill = LOOP_BG_COLOR_FILL
        requirement_cell.border = get_area_four_side_border()

    sheet2.merge_cells("C19:%s19" % get_column_letter(col+2))
    #must fisrt merge cell and set border
    set_area_four_side_border(sheet2, "C19:%s19" % get_column_letter(col+3), style='thin')

    summary_title_cell = sheet2['C19'] 
    summary_title_cell.value= "Summary - Success Rate"
    summary_title_cell.font = TAHOMA_8_BOLD_FONT
    summary_title_cell.alignment=CENTER_CENTER_ALIGNMENT

    sheet2['B19'].fill =  SUMMARY_BG_COLOR_FILL
    sheet2['B19'].border = TOP_THIN_BORDER +LEFT_THIN_BORDER
    sheet2['B20'].fill =  SUMMARY_BG_COLOR_FILL
    sheet2['B20'].border = LEFT_THIN_BORDER

    succ_rate_cell = sheet2['B21']
    succ_rate_cell.fill = SUMMARY_BG_COLOR_FILL
    succ_rate_cell.value = "Success Rate"
    succ_rate_cell.font = TAHOMA_8_BOLD_FONT
    succ_rate_cell.alignment=Alignment(wrap_text=True, horizontal='center', vertical='center')
    succ_rate_cell.border = get_area_four_side_border()

    requ_rate_cell = sheet2['B22']
    requ_rate_cell.value = "Requirement"
    requ_rate_cell.fill = SUMMARY_BG_COLOR_FILL
    requ_rate_cell.font = TAHOMA_8_BOLD_FONT
    requ_rate_cell.alignment=Alignment(wrap_text=True, horizontal='center', vertical='center')
    requ_rate_cell.border = get_area_four_side_border()

#################################################
#   Create Device Test Data Info area
#################################################
def set_Device_test_time_area(devicenum=1, start_row = 27):

    DEVICE_TITLE_ROW = start_row
    LOOP_TITLE_ROW = DEVICE_TITLE_ROW + 1
    SECTION_START_ROW = DEVICE_TITLE_ROW + 2
    AVG_COLUMN = 22

    sheet1.merge_cells("B%s:U%s" % (DEVICE_TITLE_ROW,DEVICE_TITLE_ROW))
    set_area_four_side_border(sheet1, "B%s:U%s" % (DEVICE_TITLE_ROW,DEVICE_TITLE_ROW),style='thin')
    rows = sheet1["B%s:U%s" % (DEVICE_TITLE_ROW,DEVICE_TITLE_ROW)]
    for cell in rows[0]:
        cell.fill = TITLE_BG_COLOR_FILL
    title_cell = sheet1["B%s" % DEVICE_TITLE_ROW]
    title_cell.alignment = CENTER_CENTER_ALIGNMENT
    title_cell.value =  "Device #%s - Loop Number" % devicenum
    title_cell.font = TAHOMA_8_BOLD_FONT


    for index, cell in enumerate(sheet1["B%s:U%s" % (LOOP_TITLE_ROW,LOOP_TITLE_ROW)][0], start=1):
        cell.fill = LOOP_BG_COLOR_FILL
        cell.value = index
        cell.font = TAHOMA_8_BOLD_FONT
        cell.alignment=RIGHT_CENTER_ALIGNMENT
        cell.border = get_area_four_side_border()

    pre_section_avg_title_cell=sheet1["V%s" % LOOP_TITLE_ROW]
    pre_section_avg_title_cell.value = "Average Time"
    pre_section_avg_title_cell.fill = SECTION_AVG_TIME_BG_COLOR_FILL
    pre_section_avg_title_cell.font = TAHOMA_8_BOLD_FONT
    pre_section_avg_title_cell.alignment=CENTER_CENTER_ALIGNMENT
    pre_section_avg_title_cell.border = get_area_four_side_border()

    section_cell = sheet1['A%s' % LOOP_TITLE_ROW]
    section_cell.fill = LOOP_BG_COLOR_FILL
    section_cell.value = 'Section Number'
    section_cell.font = TAHOMA_8_BOLD_FONT
    section_cell.alignment=LEFT_CENTER_ALIGNMENT
    section_cell.border = get_area_four_side_border()

    set_area_all_cell(sheet1 ,"B%s:U" % SECTION_START_ROW +str(SECTION_START_ROW+len(MAX_SECTION_LIST)-1))

    for index, section in enumerate(MAX_SECTION_LIST, start=0):
        curr_row = SECTION_START_ROW + index
        cell = sheet1.cell(row=curr_row, column=1, value=section)
        cell.fill = SECTION_BG_COLOR_FILL
        cell.font = TAHOMA_8_FONT
        cell.alignment=LEFT_CENTER_ALIGNMENT
        cell.border = get_area_four_side_border()

        for section_data in DEVICE_LIST[devicenum-1]:
            if section == section_data[0]:
                for i, module_pre_circle_time in enumerate(section_data[2], start=1):
                    cell_value = datetime.time(*map(int, module_pre_circle_time.split(":")))
                    sheet1.cell(row=curr_row, column=1+i, value=cell_value)

        avg_cell = sheet1.cell(row=curr_row, column=AVG_COLUMN, value='=SUM(B%s:U%s)/COUNTIF(B%s:U%s,">0")' % (curr_row, curr_row, curr_row, curr_row))
        avg_cell.fill = SECTION_AVG_TIME_BG_COLOR_FILL
        avg_cell.font = TAHOMA_8_FONT
        avg_cell.alignment=CENTER_CENTER_ALIGNMENT
        avg_cell.border = get_area_four_side_border()
        avg_cell.number_format = "[h]:mm:ss"

    LOOP_TIME_ROW = SECTION_START_ROW + index + 1
    LOOP_TIME_AREA = "A%s:U%s" % (LOOP_TIME_ROW, LOOP_TIME_ROW)
    for ind, cell in enumerate(sheet1[LOOP_TIME_AREA][0], start=1):
        if ind == 1:
            cell.value = "Totals"
            cell.alignment=LEFT_CENTER_ALIGNMENT
        else:
            col_letter = get_column_letter(ind)
            cell.value = '=SUM(%s%s:%s%s)' % (col_letter, SECTION_START_ROW, col_letter, LOOP_TIME_ROW-1)
            cell.alignment=RIGHT_CENTER_ALIGNMENT
            cell.number_format = "[h]:mm:ss"

        cell.fill = LOOP_TIME_BG_COLOR_FILL
        cell.font = TAHOMA_8_BOLD_FONT
        cell.border = get_area_four_side_border()

    AVG_LOOP_TIME_AREA="V%s" % LOOP_TIME_ROW
    avg_loop_time_cell = sheet1[AVG_LOOP_TIME_AREA]
    avg_loop_time_cell.value = '=SUM(%s)/COUNTIF(%s,">0")' % (LOOP_TIME_AREA, LOOP_TIME_AREA)
    avg_loop_time_cell.font = TAHOMA_8_BOLD_FONT
    avg_loop_time_cell.fill = LOOP_BG_COLOR_FILL
    avg_loop_time_cell.alignment = CENTER_CENTER_ALIGNMENT
    avg_loop_time_cell.border = get_area_four_side_border()
    avg_loop_time_cell.number_format = "[h]:mm:ss"

    return (AVG_LOOP_TIME_AREA,LOOP_TIME_ROW)

def set_Device_test_rate_area(devicenum=1, start_row = 27):
    DEVICE_TITLE_ROW = start_row
    LOOP_TITLE_ROW = DEVICE_TITLE_ROW + 1
    SECTION_START_ROW = DEVICE_TITLE_ROW + 2


    sheet2.merge_cells("D%s:W%s" % (DEVICE_TITLE_ROW,DEVICE_TITLE_ROW))
    set_area_four_side_border(sheet2, "D%s:W%s" % (DEVICE_TITLE_ROW,DEVICE_TITLE_ROW),style='thin')
    rows = sheet2["D%s:W%s" % (DEVICE_TITLE_ROW,DEVICE_TITLE_ROW)]
    for cell in rows[0]:
        cell.fill = TITLE_BG_COLOR_FILL
    title_cell = sheet2["D%s" % DEVICE_TITLE_ROW]
    title_cell.alignment = CENTER_CENTER_ALIGNMENT
    title_cell.value =  "Device #%s - Loop Number" % devicenum
    title_cell.font = TAHOMA_8_BOLD_FONT


    for index, cell in enumerate(sheet2["C%s:W%s" % (LOOP_TITLE_ROW,LOOP_TITLE_ROW)][0], start=0):
        if index == 0:
            cell.value = "Total Attempted for Every Loop"
        else:
            cell.value = "Loop %s Total Successful" % index
        cell.fill = LOOP_BG_COLOR_FILL
        cell.font = TAHOMA_8_BOLD_FONT
        cell.alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')
        cell.border = get_area_four_side_border()

    for i,str in enumerate(["Total Attempted","Total Sucessful", "Success Rate"], start=0):
        ce= sheet2.cell(row=LOOP_TITLE_ROW, column=25+i,value=str)
        if i == 2:
            ce.fill = SECTION_AVG_TIME_BG_COLOR_FILL
        else:
            ce.fill = LOOP_BG_COLOR_FILL
        ce.font = TAHOMA_8_BOLD_FONT
        ce.alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')
        ce.border = get_area_four_side_border()

    section_cell = sheet2['B%s' % LOOP_TITLE_ROW]
    section_cell.fill = LOOP_BG_COLOR_FILL
    section_cell.value = 'Section Number'
    section_cell.font = TAHOMA_8_BOLD_FONT
    section_cell.alignment=LEFT_CENTER_ALIGNMENT
    section_cell.border = get_area_four_side_border()

    for index, section in enumerate(MAX_SECTION_LIST, start=0):
        curr_row = SECTION_START_ROW + index
        cell = sheet2.cell(row=curr_row, column=2, value=section)
        cell.fill = SECTION_BG_COLOR_FILL
        cell.font = TAHOMA_8_FONT
        cell.alignment=LEFT_CENTER_ALIGNMENT
        cell.border = get_area_four_side_border()

        for section_data in DEVICE_LIST[devicenum-1]:
            if section == section_data[0]:
                sheet2.cell(row=curr_row, column=3, value=section_data[1])
                for i, module_pre_circle_sucetimes in enumerate(section_data[3], start=1):
                    sheet2.cell(row=curr_row, column=3+i, value=module_pre_circle_sucetimes)

        t_a_cell = sheet2.cell(row=curr_row, column=25, value='=PRODUCT(C%s,COUNT(D%s:W%s))' % (curr_row, curr_row, curr_row))
        #t_a_cell.fill = SECTION_AVG_TIME_BG_COLOR_FILL
        t_a_cell.font = TAHOMA_8_FONT
        t_a_cell.alignment=CENTER_CENTER_ALIGNMENT
        t_a_cell.border = get_area_four_side_border()

        t_s_cell = sheet2.cell(row=curr_row, column=26, value='=SUM(D%s:W%s)' % (curr_row, curr_row))
        #t_s_cell.fill = SECTION_AVG_TIME_BG_COLOR_FILL
        t_s_cell.font = TAHOMA_8_FONT
        t_s_cell.alignment=CENTER_CENTER_ALIGNMENT
        t_s_cell.border = get_area_four_side_border()

        s_r_cell = sheet2.cell(row=curr_row, column=27, value='=(Z%s/Y%s)*100' % (curr_row, curr_row))
        s_r_cell.fill = SECTION_AVG_TIME_BG_COLOR_FILL
        s_r_cell.font = TAHOMA_8_FONT
        s_r_cell.alignment=CENTER_CENTER_ALIGNMENT
        s_r_cell.border = get_area_four_side_border()

    LOOP_TIME_ROW = SECTION_START_ROW + index + 1
    LOOP_TIME_AREA = "B%s:Z%s" % (LOOP_TIME_ROW, LOOP_TIME_ROW)
    for ind, cell in enumerate(sheet2[LOOP_TIME_AREA][0], start=1):
        if ind == 1:
            cell.value = "Totals"
            cell.alignment=LEFT_CENTER_ALIGNMENT
        else:
            col_letter = get_column_letter(ind+1)
            if col_letter == 'X':
                continue
            cell.value = '=SUM(%s%s:%s%s)' % (col_letter, SECTION_START_ROW, col_letter, LOOP_TIME_ROW-1)
            cell.alignment=RIGHT_CENTER_ALIGNMENT
            #cell.number_format = "[h]:mm:ss"

        cell.fill = LOOP_TIME_BG_COLOR_FILL
        cell.font = TAHOMA_8_BOLD_FONT
        cell.border = get_area_four_side_border()

    AVG_LOOP_TIME_AREA="AA%s" % LOOP_TIME_ROW
    avg_loop_time_cell = sheet2[AVG_LOOP_TIME_AREA]
    avg_loop_time_cell.value = '=(Z%s/Y%s)*100' % (LOOP_TIME_ROW, LOOP_TIME_ROW)
    avg_loop_time_cell.font = TAHOMA_8_BOLD_FONT
    avg_loop_time_cell.fill = SECTION_AVG_TIME_BG_COLOR_FILL
    avg_loop_time_cell.alignment = CENTER_CENTER_ALIGNMENT
    avg_loop_time_cell.border = get_area_four_side_border()

    set_area_all_cell_2(sheet2 ,"C%s:W%s" % (SECTION_START_ROW ,SECTION_START_ROW+len(MAX_SECTION_LIST)-1))

    return (AVG_LOOP_TIME_AREA,LOOP_TIME_ROW)


def test_time_summary(sheet, Devices, start_row):
    for devicenum in xrange(Devices):
        a, start_row = set_Device_test_time_area(devicenum+1, start_row+3)
        sheet["%s21" % get_column_letter(devicenum+3)] = "=%s" % a

    return start_row

def test_rate_summary(sheet, Devices,start_row):
    s_row = start_row
    for devicenum in xrange(Devices):
        a, s_row = set_Device_test_rate_area(devicenum+1, s_row+3)

    l = len(MAX_SECTION_LIST)
    for section in xrange(l):
        b=()
        section_start_row = section + start_row + 5
        for devicenum in xrange(Devices):
            b = b + (section_start_row + (l+5)*devicenum,)
        Z_map = reduce(lambda x,y: x+"+"+y , map(lambda x: "Z"+str(x),b))
        Y_map = reduce(lambda x,y: x+"+"+y , map(lambda x: "Y"+str(x),b))
        sheet["%s21" % get_column_letter(section+3)] = "=SUM(%s)/SUM(%s)*100" % (Z_map, Y_map)

    return s_row

def set_Mutil_Device_test_data_area(sheet, set_handler):
    device_test_data_start_row=24
    evice_test_data_end_row = set_handler(sheet, DEVICENUM, device_test_data_start_row)
    sheet_end_row = evice_test_data_end_row+10
    end_area = "A%s:V%s" % (sheet_end_row,sheet_end_row)
    sheet.merge_cells(end_area)
    set_area_four_side_border(sheet, end_area)
    for cell in sheet[end_area][0]:
        cell.fill = SHEET_END_BG_COLOR_FILL
    sheet[end_area.split(":")[0]].value = 'AT&T Proprietary - (RESTRICTED)\nOnly for use by authorized individuals\nwithin the AT&T companies and not for general distribution'
    sheet[end_area.split(":")[0]].alignment=Alignment(wrap_text=True,horizontal='center', vertical='center') 
    sheet[end_area.split(":")[0]].font = ARIAL_10_FONT
    sheet.row_dimensions[sheet_end_row].height = 72

def create_xlsx_file():
    #sheet 1
    SHEET1_TABLE_INSTRCUCTIONS_STR = [ 
                            "Table Instructions:", 
                            "1. OEMs will need to log and record the time taken to execute each stability section for each loop for each device",
                            "2. Leave the columns for loops not executed at 0:00:00",
                            "3. Time is formatted in Hours:Minutes:Seconds",
                            ]

    SHEET2_SHEET1_TABLE_INSTRCUCTIONS_STR = [
                            "Table Instructions:",
                            "1. OEMs will need to fill in the number of events planned to be executed for eack stability test section colmun C (Total Attempted for Every Loop)",
                            "2. OEMs will need to log and record the number of events completed sucessfully for each loop in each section for each device (1-7)",
                            "3. Leave the columns for loops not executed blank",
                            "4. Example for Device #1 is below (numbers in example are not actual counts or even close)",
                            ]

    set_Rights_area(sheet1)
    set_Table_instructions_area(sheet1, SHEET1_TABLE_INSTRCUCTIONS_STR)
    set_Loop_time_summary_area()
    set_Mutil_Device_test_data_area(sheet1, test_time_summary)

    #sheet 2 
    set_Rights_area(sheet2)
    set_Table_instructions_area(sheet2, SHEET2_SHEET1_TABLE_INSTRCUCTIONS_STR)
    set_Success_rate_summary_area()
    set_Mutil_Device_test_data_area(sheet2, test_rate_summary)

    wb.save(FILENAME)

def write_data_to_xlsx_file(*jsonfiles):
    global MAX_SECTION_LIST
    global DEVICE_LIST
    global DEVICENUM
    for ind , jsonfile in enumerate(jsonfiles, start=1):
        LJson.JsonLoadFromFile(jsonfile)
        data = LJson.ToJson()
        modu_data = data["ModuleData"]

        DEVICE_LIST.append([])
        section_list = DEVICE_LIST[-1]

        for pre_modu_item in modu_data:
            if pre_modu_item['PerCricleSucessTimes']:
                section_name = "5.1." + str(ALL_SECTION_LIST.index(pre_modu_item['ModuleName']))+'_'+pre_modu_item['ModuleName']
                section_list.append((section_name, pre_modu_item['TotalTimes'], pre_modu_item['PerCricleRunTime'], pre_modu_item['PerCricleSucessTimes']))
                if section_name not in MAX_SECTION_LIST:
                    MAX_SECTION_LIST.append(section_name)
        if len(section_list) <= 0:
            DEVICE_LIST.pop()

    DEVICENUM = len(DEVICE_LIST)

    create_xlsx_file()